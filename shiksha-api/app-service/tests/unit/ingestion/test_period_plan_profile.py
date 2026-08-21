"""Unit test for the period_plan_chapter_grouping omni-ingest profile.

Builds a tiny xlsx (two chapters, one vertically-merged chapter_number cell),
runs it through the local omni-ingest path dependency using only the xlsx-read
+ transform steps (no LLM, runs offline), and asserts correct chapter grouping
and index_path construction.

The test is deliberately synchronous (no pytest-asyncio marker): it drives the
async pipeline with a single asyncio.run() so it runs identically regardless of
the pytest-asyncio asyncio_mode in effect.
"""

import asyncio
import io
import json
import tempfile
from pathlib import Path

import openpyxl

from omni_ingest.core.pipeline import IngestionContext, create_pipeline_from_config
from omni_ingest.port.metadata_store import NullMetadataStore

PROFILE = Path(__file__).resolve().parents[3] / "ingestion" / "profiles" / "period_plan_chapter_grouping.yaml"


def _make_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grade 6"
    # Mixed case and trailing whitespace, matching real Bloom workbooks
    # (headers are not consistent across grade sheets in production).
    ws.append(["Chapter Number", "Chapter title", "Period ", "Topic", "Learning Outcomes"])
    ws.append([1, "Knowing Our Numbers", "P1", "Large Numbers", "LO1"])
    ws.append([None, None, "P1", "Estimation", "LO2"])  # same period, second topic
    ws.append([None, None, "P2", "Roman Numerals", "LO3"])  # chapter_number/title merged from row 2
    ws.append([2, "Whole Numbers", "P1", "Number Line", "LO4"])
    # Separate vertical merges (one column each) forward-filled by the parser --
    # matches how the Bloom workbook merges chapter_number and chapter_title down.
    ws.merge_cells("A2:A4")
    ws.merge_cells("B2:B4")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


_WORKFLOW_TEMPLATE = {
    "_id": "social-lesson-plan",
    "sections": [
        {"id": "lesson_mind_mapping", "description": "Mind map for: ${READING_CONTENT}"},
        {
            "id": "lesson_discourse",
            "description": (
                "${READING_CONTENT}\nMajor: ${DISCOURSE_TYPES_MAJOR}\nMinor: ${DISCOURSE_TYPES}"
            ),
        },
    ],
}


def _run_profile(data: bytes, grade: str, subject: str, **extra_config):
    async def _run():
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = Path(tmp) / "periods.xlsx"
            xlsx_path.write_bytes(data)
            config_args = {"grade": grade, "subject": subject, "workflow_template": json.dumps(_WORKFLOW_TEMPLATE), **extra_config}
            runner = create_pipeline_from_config(PROFILE, config_args=config_args)
            ctx = IngestionContext(resource=xlsx_path, store=NullMetadataStore(), domain_profile="period-plan")
            results = await runner.run(ctx)
            failures = [r.error for r in results if r.status.value == "failure"]
            assert not failures, failures
            return ctx.metadata

    return asyncio.run(_run())


def test_period_plan_chapter_grouping():
    chapters = _run_profile(_make_xlsx(), grade="Grade 6", subject="math")["chapters"]

    assert isinstance(chapters, list)
    assert len(chapters) == 2, f"expected 2 chapters, got {len(chapters)}"

    by_number = {c["chapter_number"]: c for c in chapters}

    first = by_number[1]
    assert first["chapter_title"] == "Knowing Our Numbers"
    assert first["index_path"] == "qdrant/SCERT/chapter_id:Medium=english,Grade=Grade 6,Subject=math,Number=1"
    # P1 has two topic rows (merged chapter_number/title forward-filled onto both);
    # P2 has one. Two periods -> two topic_groups.
    assert len(first["topic_groups"]) == 2
    p1, p2 = first["topic_groups"]
    assert p1["topic"] == ["Large Numbers", "Estimation"]
    assert p1["learning_outcomes"] == ["LO1", "LO2"]
    assert p2["topic"] == ["Roman Numerals"]
    assert p2["learning_outcomes"] == ["LO3"]

    second = by_number[2]
    assert second["chapter_title"] == "Whole Numbers"
    assert second["index_path"] == "qdrant/SCERT/chapter_id:Medium=english,Grade=Grade 6,Subject=math,Number=2"
    assert len(second["topic_groups"]) == 1
    assert second["topic_groups"][0]["topic"] == ["Number Line"]
    assert second["topic_groups"][0]["learning_outcomes"] == ["LO4"]


def test_period_plan_payload_building():
    payloads = _run_profile(_make_xlsx(), grade="Grade 6", subject="math")["payloads"]

    # 2 topic_groups in chapter 1 + 1 in chapter 2 = 3 payloads
    assert len(payloads) == 3
    by_doc_id = {p["_id"]: p for p in payloads}

    first_period = by_doc_id[
        "Board=BSE-TG/Medium=english/Grade=Grade 6/Subject=math/Number=1/Level=SUBTOPIC/Topics=Large Numbers & Estimation"
    ]
    assert first_period["chapter_id"] == "Board=BSE-TG,Medium=english,Grade=Grade 6,Subject=math,Number=1,Title=Knowing Our Numbers"
    assert first_period["user_id"] == "ADMIN"
    assert first_period["workflow_id"] == "social-lesson-plan"
    assert first_period["lp_level"] == "SUBTOPIC"
    assert first_period["subtopics"] == ["Large Numbers", "Estimation"]
    assert first_period["learning_outcomes"] == ["LO1", "LO2"]
    assert first_period["chapter_info"] == {
        "id": "Board=BSE-TG,Medium=english,Grade=Grade 6,Subject=math,Number=1,Title=Knowing Our Numbers",
        "chapter_title": "Knowing Our Numbers",
        "index_path": "qdrant/SCERT/chapter_id:Medium=english,Grade=Grade 6,Subject=math,Number=1",
    }
    # first topic_group of a chapter (topic_counter == 1) keeps lesson_mind_mapping
    section_ids = [s["id"] for s in first_period["workflow"]["sections"]]
    assert section_ids == ["lesson_mind_mapping", "lesson_discourse"]
    discourse = next(s for s in first_period["workflow"]["sections"] if s["id"] == "lesson_discourse")
    assert "CHAPTER: Knowing Our Numbers" in discourse["description"]
    assert "SUBTOPICS: Large Numbers & Estimation" in discourse["description"]
    assert "LEARNING OUTCOMES: LO1, LO2" in discourse["description"]
    assert "Major: Lab Report; Project Work; Field Trip Report; Model Making" in discourse["description"]
    assert "Minor: Concept Map; Flow Chart; Diagram; Data Table; Quiz" in discourse["description"]
    assert "${" not in discourse["description"]

    second_period = by_doc_id[
        "Board=BSE-TG/Medium=english/Grade=Grade 6/Subject=math/Number=1/Level=SUBTOPIC/Topics=Roman Numerals"
    ]
    assert second_period["subtopics"] == ["Roman Numerals"]
    # second topic_group of the same chapter (topic_counter == 2) drops lesson_mind_mapping
    assert [s["id"] for s in second_period["workflow"]["sections"]] == ["lesson_discourse"]

    chapter2_period = by_doc_id[
        "Board=BSE-TG/Medium=english/Grade=Grade 6/Subject=math/Number=2/Level=SUBTOPIC/Topics=Number Line"
    ]
    assert chapter2_period["subtopics"] == ["Number Line"]
    assert [s["id"] for s in chapter2_period["workflow"]["sections"]] == ["lesson_mind_mapping", "lesson_discourse"]
